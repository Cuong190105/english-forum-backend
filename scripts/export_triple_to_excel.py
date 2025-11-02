from __future__ import annotations
import csv
import json
import os
import re
from pathlib import Path
import sys
from typing import List, Dict, Any, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except Exception:
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl")


def _append_csv(ws: Worksheet, csv_path: Path) -> int:
    if not csv_path.exists():
        return 0
    with csv_path.open('r', encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        n = 0
        for row in reader:
            ws.append(row)
            n += 1
        return n


def _autosize(ws: Worksheet) -> None:
    try:
        for col_idx, _ in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            col_letter = get_column_letter(col_idx)
            width = 0
            for cell in ws[col_letter]:
                v = cell.value
                if v is None:
                    continue
                l = len(str(v))
                if l > width:
                    width = l
            ws.column_dimensions[col_letter].width = min(max(width + 2, 10), 100)
    except Exception:
        pass


# Ensure project root is importable when running as a script
def _ensure_repo_root_on_path() -> None:
    try:
        here = Path(__file__).resolve()
        repo_root = here.parent.parent  # scripts/ -> project root
        if str(repo_root) not in sys.path:
            sys.path.insert(0, str(repo_root))
    except Exception:
        pass


_ensure_repo_root_on_path()

# ===== Pre-export enrichment: recompute structure/semantics and paired deltas =====
def _safe_topic(topic: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", topic or '')


def _load_csv_dicts(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open('r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f, skipinitialspace=True)
        if reader.fieldnames:
            reader.fieldnames = [(fn or '').strip() for fn in reader.fieldnames]
        for r in reader:
            if r is None:
                continue
            rows.append({(k or '').strip(): (v.strip() if isinstance(v, str) else v) for k, v in r.items()})
    return rows


def _verdict_to_score(hw_type: str, verdict: str) -> float:
    v = (verdict or '').strip().lower()
    if hw_type == 'mcq':
        return {'correct':1.0, 'ambiguous':0.5, 'incorrect':0.0}.get(v, 0.0)
    else:
        return {'acceptable':1.0, 'unacceptable':0.0}.get(v, 0.0)


def _precompute_triple(run_dir: Path) -> None:
    """Recompute structural/semantic metrics into per_item.csv and produce summary.csv,
    paired_overall.csv, and winloss.csv so Excel export has complete data.
    """
    from benchmark.score import mcq_semantics
    from benchmark.validate import validate_items
    from benchmark.report import write_per_item_csv, write_summary_csv, write_paired_overall_csv, write_winloss_csv
    from scripts.recompute_reports import compute_group_stats

    per_item_csv = run_dir / 'per_item.csv'
    rows = _load_csv_dicts(per_item_csv)
    if not rows:
        print(f"[precompute] No rows loaded from {per_item_csv}; skipping enrichment.")
        return

    # Group rows so we can load pred/gold once per (config, topic, type, seed, sha)
    groups: Dict[Tuple[str,str,str,str,str], List[int]] = {}
    for i, r in enumerate(rows):
        key = (
            str(r.get('config','')).strip(),
            str(r.get('topic','')).strip(),
            str(r.get('type','')).strip(),
            str(r.get('seed','')).strip(),
            str(r.get('source_text_sha','')).strip(),
        )
        groups.setdefault(key, []).append(i)

    sem_fail_count = 0
    updated_sem_count = 0
    for (config, topic, hw_type, seed, sha), idxs in groups.items():
        if not (config and topic and hw_type and seed and sha):
            continue
        st = _safe_topic(topic)
        pred_path = Path(f"benchmark/pred/{config}/{st}/{hw_type}/{sha}/seed{seed}.json")
        gold_path = Path(f"benchmark/gold/{st}/{hw_type}/{sha}/seed0.json")
        try:
            pred = json.loads(pred_path.read_text(encoding='utf-8')) if pred_path.exists() else []
        except Exception as e:
            print(f"[precompute] Failed reading {pred_path}: {e}")
            pred = []
        try:
            gold = json.loads(gold_path.read_text(encoding='utf-8')) if gold_path.exists() else []
        except Exception as e:
            print(f"[precompute] Failed reading {gold_path}: {e}")
            gold = []

        # Structural validation: collect invalid item indices (1-based) from pred
        ok_pred, errs_pred, _ = validate_items(pred, hw_type, expected_count=None)
        invalid_idx_pred = set()
        for e in errs_pred:
            if isinstance(e, str) and e.startswith('i') and ':' in e:
                try:
                    idx_str = e[1:e.index(':')]
                    ii = int(idx_str)
                    invalid_idx_pred.add(ii)
                except Exception:
                    pass

        for i in idxs:
            r = rows[i]
            try:
                idx = int(str(r.get('idx','')).strip() or '0')
            except Exception:
                continue
            if idx <= 0:
                continue
            pi = pred[idx-1] if idx-1 < len(pred) else {}
            gi = gold[idx-1] if idx-1 < len(gold) else {}

            # structural_valid (1 if passes, else 0)
            r['structural_valid'] = 0 if (idx in invalid_idx_pred) else 1

            # semantics
            if hw_type == 'mcq' and pi and gi:
                try:
                    ps, ans, div, item = mcq_semantics(pi, gi)
                    r['question_sim'] = round(ps, 4)
                    r['ans_sim'] = ('' if ans is None else round(ans, 4))
                    r['distractor_diversity'] = round(div, 4)
                    r['ans_score'] = ('' if ans is None else round(ans, 4))
                    r['item_score'] = round(item, 4)
                    updated_sem_count += 1
                except Exception as e:
                    # Embedding/API issues shouldn't abort the whole export; record and continue
                    if sem_fail_count < 3:
                        print(f"[precompute] semantics failed for (cfg={config}, topic={topic}, idx={idx}): {e}")
                    sem_fail_count += 1
            else:
                # Ignore non-MCQ types entirely (no fill support)
                pass

            # Ensure legacy primary judge fields populated (use Gemini by default)
            if not str(r.get('judge_verdict','')).strip():
                r['judge_verdict'] = r.get('judge_gemini_verdict','')
            if str(r.get('judge_score','')).strip() == '':
                try:
                    r['judge_score'] = round(float(_verdict_to_score(hw_type, r.get('judge_verdict',''))), 4)
                except Exception:
                    r['judge_score'] = ''
            rows[i] = r

    # Write back per_item
    write_per_item_csv(per_item_csv, rows)

    print(f"[precompute] structural updated for ~{len(rows)} rows; semantics updated for {updated_sem_count} rows; failures={sem_fail_count}")

    # summary.csv
    summary_rows = compute_group_stats(rows)
    write_summary_csv(run_dir / 'summary.csv', summary_rows)

    # paired_overall.csv and winloss.csv
    # Build deltas between minimal and cot per (sha, topic, type, seed)
    groups2: Dict[Tuple[str,str,str,str], Dict[str, Dict[str, float]]] = {}
    for r in rows:
        sha = str(r.get('source_text_sha','')).strip()
        topic = str(r.get('topic','')).strip()
        t = str(r.get('type','')).strip()
        if t != 'mcq':
            continue
        seed = str(r.get('seed','')).strip()
        cfg = str(r.get('config','')).strip()
        if not (sha and topic and t and seed and cfg):
            continue
        try:
            item_score = float(r.get('item_score','')) if str(r.get('item_score','')).strip() != '' else None
            judge_score = float(r.get('judge_score','')) if str(r.get('judge_score','')).strip() != '' else None
        except Exception:
            item_score = None
            judge_score = None
        key = (sha, topic, t, seed)
        groups2.setdefault(key, {})
        if cfg not in groups2[key]:
            groups2[key][cfg] = {'semantic_sum':0.0, 'semantic_n':0, 'judge_sum':0.0, 'judge_n':0}
        if item_score is not None:
            groups2[key][cfg]['semantic_sum'] += item_score
            groups2[key][cfg]['semantic_n'] += 1
        if judge_score is not None:
            groups2[key][cfg]['judge_sum'] += judge_score
            groups2[key][cfg]['judge_n'] += 1

    from statistics import mean, stdev
    from math import sqrt
    deltas_sem: List[float] = []
    deltas_j: List[float] = []
    s_w = s_l = s_t = 0
    j_w = j_l = j_t = 0
    for _, cfgmap in groups2.items():
        if 'minimal' in cfgmap and 'cot' in cfgmap:
            def avg(d: Dict[str,float], s: str, n: str) -> float:
                return (d.get(s,0.0) / d.get(n,1)) if d.get(n,0) > 0 else 0.0
            min_sem = avg(cfgmap['minimal'], 'semantic_sum', 'semantic_n')
            cot_sem = avg(cfgmap['cot'], 'semantic_sum', 'semantic_n')
            min_j = avg(cfgmap['minimal'], 'judge_sum', 'judge_n')
            cot_j = avg(cfgmap['cot'], 'judge_sum', 'judge_n')
            d_sem = cot_sem - min_sem
            d_j = cot_j - min_j
            deltas_sem.append(d_sem)
            deltas_j.append(d_j)
            eps = 1e-12
            if d_sem > eps:
                s_w += 1
            elif d_sem < -eps:
                s_l += 1
            else:
                s_t += 1
            if d_j > eps:
                j_w += 1
            elif d_j < -eps:
                j_l += 1
            else:
                # tie -> break by semantic
                if d_sem > 1e-6:
                    j_w += 1
                elif d_sem < -1e-6:
                    j_l += 1
                else:
                    j_t += 1

    def ci_stats(arr: List[float]) -> Dict[str, float]:
        if not arr:
            return dict(mean=0.0, std=0.0, se=0.0, low=0.0, high=0.0)
        m = mean(arr)
        sd = stdev(arr) if len(arr) > 1 else 0.0
        se = (sd / sqrt(len(arr))) if len(arr) > 1 else 0.0
        low = max(-1.0, m - 1.96 * se)
        high = min(1.0, m + 1.96 * se)
        return dict(mean=m, std=sd, se=se, low=low, high=high)

    s_stats = ci_stats(deltas_sem)
    j_stats = ci_stats(deltas_j)
    s_H = max(0.0, (s_stats['high'] - s_stats['low'])/2.0)
    j_H = max(0.0, (j_stats['high'] - j_stats['low'])/2.0)
    paired_rows = [
        {
            'run_id': rows[0].get('run_id','') if rows else '',
            'metric': 'semantic',
            'n_pairs': len(deltas_sem),
            'delta_mean': round(s_stats['mean'], 4),
            'delta_std': round(s_stats['std'], 4),
            'delta_se': round(s_stats['se'], 4),
            'delta_ci95_low': round(s_stats['low'], 4),
            'delta_ci95_high': round(s_stats['high'], 4),
            'delta_ci95_halfwidth_H': round(s_H, 4),
            'delta_ci_level': 0.95,
            'delta_H_target': 0.04,
            'delta_meets_H_target': 1 if s_H <= 0.04 else 0,
            'delta_significant': 1 if (s_stats['low'] > 0.0 or s_stats['high'] < 0.0) else 0,
        },
        {
            'run_id': rows[0].get('run_id','') if rows else '',
            'metric': 'judge',
            'n_pairs': len(deltas_j),
            'delta_mean': round(j_stats['mean'], 4),
            'delta_std': round(j_stats['std'], 4),
            'delta_se': round(j_stats['se'], 4),
            'delta_ci95_low': round(j_stats['low'], 4),
            'delta_ci95_high': round(j_stats['high'], 4),
            'delta_ci95_halfwidth_H': round(j_H, 4),
            'delta_ci_level': 0.95,
            'delta_H_target': 0.05,
            'delta_meets_H_target': 1 if j_H <= 0.05 else 0,
            'delta_significant': 1 if (j_stats['low'] > 0.0 or j_stats['high'] < 0.0) else 0,
        },
    ]
    write_paired_overall_csv(run_dir / 'paired_overall.csv', paired_rows)

    # Win/loss summary
    def _binom_p_two_sided(wins: int, losses: int) -> float:
        import math
        n = wins + losses
        if n == 0:
            return 1.0
        k = wins
        def pmf(i: int) -> float:
            return math.comb(n, i) * (0.5 ** n)
        def cdf(i: int) -> float:
            return sum(pmf(j) for j in range(0, i+1))
        c_low = cdf(k)
        c_high = 1.0 - cdf(k-1) if k > 0 else 1.0
        p = 2.0 * min(c_low, c_high)
        return min(max(p, 0.0), 1.0)

    s_n_eff = s_w + s_l
    j_n_eff = j_w + j_l
    s_rate = (s_w / s_n_eff) if s_n_eff > 0 else 0.0
    j_rate = (j_w / j_n_eff) if j_n_eff > 0 else 0.0
    s_p = _binom_p_two_sided(s_w, s_l)
    j_p = _binom_p_two_sided(j_w, j_l)
    winloss_rows = [{
        'run_id': rows[0].get('run_id','') if rows else '',
        'semantic_wins': s_w,
        'semantic_losses': s_l,
        'semantic_ties': s_t,
        'semantic_n_effective': s_n_eff,
        'semantic_win_rate': round(s_rate, 4),
        'semantic_binomial_p': round(s_p, 6),
        'judge_wins': j_w,
        'judge_losses': j_l,
        'judge_ties': j_t,
        'judge_n_effective': j_n_eff,
        'judge_win_rate': round(j_rate, 4),
        'judge_binomial_p': round(j_p, 6),
    }]
    write_winloss_csv(run_dir / 'winloss.csv', winloss_rows)


def build_excel(run_dir: Path) -> Path:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'summary'  # generic summary with structural/semantic/judge aggregates
    ws2 = wb.create_sheet('summary_gemini')
    ws3 = wb.create_sheet('summary_claude')
    ws4 = wb.create_sheet('summary_deepseek')
    ws5 = wb.create_sheet('inter_judge_three')
    ws5m = wb.create_sheet('inter_judge_three_multi')
    ws6 = wb.create_sheet('paired_overall')
    ws7 = wb.create_sheet('winloss')
    ws8 = wb.create_sheet('per_item')

    # Append CSVs if present
    n_sum = _append_csv(ws1, run_dir / 'summary.csv')
    n_g = _append_csv(ws2, run_dir / 'summary_gemini.csv')
    n_c = _append_csv(ws3, run_dir / 'summary_claude.csv')
    n_d = _append_csv(ws4, run_dir / 'summary_deepseek.csv')
    n_ijt = _append_csv(ws5, run_dir / 'inter_judge_by_topic_three.csv')
    n_ijm = _append_csv(ws5m, run_dir / 'inter_judge_by_topic_three_multi.csv')
    n_po = _append_csv(ws6, run_dir / 'paired_overall.csv')
    n_wl = _append_csv(ws7, run_dir / 'winloss.csv')
    n_pi = _append_csv(ws8, run_dir / 'per_item.csv')

    print(
        f"Rows: summary={n_sum}, summary_gemini={n_g}, summary_claude={n_c}, summary_deepseek={n_d}, "
        f"inter_three={n_ijt}, inter_three_multi={n_ijm}, paired_overall={n_po}, winloss={n_wl}, per_item={n_pi}"
    )

    for ws in (ws1, ws2, ws3, ws4, ws5, ws5m, ws6, ws7, ws8):
        _autosize(ws)

    out_path = run_dir / 'report_triple.xlsx'
    if out_path.exists():
        try:
            wb.save(out_path)
        except PermissionError:
            from datetime import datetime
            backup = run_dir / f'report_triple_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            print(f"Warning: {out_path} locked, saving to {backup}")
            wb.save(backup)
            out_path = backup
    else:
        wb.save(out_path)
    return out_path


def main():
    import argparse
    ap = argparse.ArgumentParser(description='Export triple-judge CSVs to an Excel workbook.')
    ap.add_argument('--run-id', required=True, help='Report run id (folder under benchmark/reports)')
    ap.add_argument('--open-file', action='store_true', help='Open the Excel file after export (Windows)')
    args = ap.parse_args()

    run_dir = Path(f"benchmark/reports/{args.run_id}")
    # Pre-export: ensure per_item has structure/semantics and paired_overall/winloss exist
    try:
        _precompute_triple(run_dir)
    except Exception as e:
        print(f"Warning: precompute step failed: {e}")
    out = build_excel(run_dir)
    print(f"Excel written: {out}")
    if args.open_file:
        try:
            os.startfile(out)
        except Exception:
            pass


if __name__ == '__main__':
    main()
