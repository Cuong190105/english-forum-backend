"""Restore DeepSeek judge2 data from report3.xlsx to the reports."""
from __future__ import annotations
import sys
import csv
import os
from pathlib import Path
from typing import List, Dict, Any

# Fix Unicode encoding for Windows
os.environ['PYTHONIOENCODING'] = 'utf-8'

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl


def load_deepseek_data_from_report3() -> Dict[tuple, Dict[str, Any]]:
    """Load DeepSeek judge2 data from report3.xlsx."""
    report3_path = Path("benchmark/reports/20251029_161124/report3.xlsx")
    
    if not report3_path.exists():
        raise FileNotFoundError(f"report3.xlsx not found: {report3_path}")
    
    wb = openpyxl.load_workbook(report3_path, data_only=True)
    
    # Use "compare" sheet which has per_item data with judge2
    if 'compare' in wb.sheetnames:
        per_item_sheet = wb['compare']
    else:
        # Find per_item sheet
        per_item_sheet = None
        for sheet_name in wb.sheetnames:
            if 'per_item' in sheet_name.lower() or 'item' in sheet_name.lower():
                per_item_sheet = wb[sheet_name]
                break
        
        if not per_item_sheet:
            # Try first sheet
            per_item_sheet = wb[wb.sheetnames[0]]
    
    print(f"Using sheet: {per_item_sheet.title}")
    
    # Read header (first row)
    headers = []
    for col in range(1, per_item_sheet.max_column + 1):
        cell_value = per_item_sheet.cell(1, col).value
        headers.append(str(cell_value) if cell_value else '')
    
    # Find column indices
    # Find column indices
    config_idx = next((i for i, h in enumerate(headers, 1) if str(h).lower() == 'config'), None)
    topic_idx = next((i for i, h in enumerate(headers, 1) if str(h).lower() == 'topic'), None)
    type_idx = next((i for i, h in enumerate(headers, 1) if str(h).lower() == 'type'), None)
    seed_idx = next((i for i, h in enumerate(headers, 1) if str(h).lower() == 'seed'), None)
    sha_idx = next((i for i, h in enumerate(headers, 1) if 'source_text' in str(h).lower() and 'sha' in str(h).lower()), None)
    source_text_idx = next((i for i, h in enumerate(headers, 1) if str(h).lower() == 'source_text'), None)
    idx_idx = next((i for i, h in enumerate(headers, 1) if str(h).lower() == 'idx'), None)
    j2_verdict_idx = next((i for i, h in enumerate(headers, 1) if 'judge2_verdict' in str(h).lower()), None)
    j2_score_idx = next((i for i, h in enumerate(headers, 1) if 'judge2_score' in str(h).lower()), None)
    j2_why_idx = next((i for i, h in enumerate(headers, 1) if 'judge2_why' in str(h).lower()), None)
    
    if any(x is None for x in [config_idx, topic_idx, type_idx, seed_idx, idx_idx, j2_verdict_idx, j2_score_idx]):
        raise ValueError(f"Missing required columns. Found: {headers}")
    
    # If no sha_idx, we'll use idx-based matching only
    if not sha_idx and not source_text_idx:
        print("WARNING: No source_text_sha or source_text column found. Will match by (config, topic, type, seed, idx) only.")
    
    deepseek_map = {}
    
    # Read rows (skip header, start from row 2)
    for row_idx in range(2, per_item_sheet.max_row + 1):
        config = str(per_item_sheet.cell(row_idx, config_idx).value or '').strip()
        topic = str(per_item_sheet.cell(row_idx, topic_idx).value or '').strip()
        hw_type = str(per_item_sheet.cell(row_idx, type_idx).value or '').strip()
        seed = str(per_item_sheet.cell(row_idx, seed_idx).value or '').strip()
        idx = str(per_item_sheet.cell(row_idx, idx_idx).value or '').strip()
        
        # Get sha if available
        if sha_idx:
            sha_cell = per_item_sheet.cell(row_idx, sha_idx)
            sha = str(sha_cell.value or '').strip() if sha_cell.value else ''
        elif source_text_idx:
            # Hash source_text if available
            source_text_cell = per_item_sheet.cell(row_idx, source_text_idx)
            import hashlib
            source_text = str(source_text_cell.value or '').strip() if source_text_cell.value else ''
            sha = hashlib.sha256(source_text.encode('utf-8')).hexdigest() if source_text else ''
        else:
            sha = ''  # Will match by config, topic, type, seed, idx only
        
        if not all([config, topic, hw_type, seed, idx]):
            continue
        
        # Key: use sha if available, otherwise match without it
        if sha:
            key = (config, topic, hw_type, seed, sha, idx)
        else:
            key = (config, topic, hw_type, seed, idx)
        
        j2_verdict = str(per_item_sheet.cell(row_idx, j2_verdict_idx).value or '').strip() if j2_verdict_idx else ''
        j2_score = str(per_item_sheet.cell(row_idx, j2_score_idx).value or '').strip() if j2_score_idx else ''
        j2_why_cell = per_item_sheet.cell(row_idx, j2_why_idx) if j2_why_idx else None
        j2_why = str(j2_why_cell.value or '').strip() if j2_why_cell else ''
        
        # Only store if not empty and not same as judge1 (to identify DeepSeek data)
        if j2_verdict and j2_verdict.lower() != 'error':
            deepseek_map[key] = {
                'verdict': j2_verdict,
                'score': j2_score,
                'why': j2_why,
            }
    
    print(f"Loaded {len(deepseek_map)} DeepSeek judge2 entries from report3.xlsx")
    return deepseek_map


def restore_deepseek_to_report(run_id: str, deepseek_map: Dict[tuple, Dict[str, Any]]):
    """Restore DeepSeek judge2 data to a report."""
    run_dir = Path(f"benchmark/reports/{run_id}")
    per_item_csv = run_dir / 'per_item.csv'
    
    if not per_item_csv.exists():
        raise FileNotFoundError(f"per_item.csv not found: {per_item_csv}")
    
    # Load current per_item.csv
    rows = []
    with per_item_csv.open('r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    
    print(f"Loaded {len(rows)} rows from {per_item_csv}")
    
    # Update rows with DeepSeek data
    updated_count = 0
    for row in rows:
        sha = row.get('source_text_sha', '').strip()
        
        # Try matching with sha first, then without
        if sha:
            key_with_sha = (
                row.get('config', '').strip(),
                row.get('topic', '').strip(),
                row.get('type', '').strip(),
                row.get('seed', '').strip(),
                sha,
                row.get('idx', '').strip(),
            )
            deepseek_data = deepseek_map.get(key_with_sha)
            
            # If not found, try without sha
            if not deepseek_data:
                key_no_sha = (
                    row.get('config', '').strip(),
                    row.get('topic', '').strip(),
                    row.get('type', '').strip(),
                    row.get('seed', '').strip(),
                    row.get('idx', '').strip(),
                )
                deepseek_data = deepseek_map.get(key_no_sha)
        else:
            key_no_sha = (
                row.get('config', '').strip(),
                row.get('topic', '').strip(),
                row.get('type', '').strip(),
                row.get('seed', '').strip(),
                row.get('idx', '').strip(),
            )
            deepseek_data = deepseek_map.get(key_no_sha)
        if deepseek_data:
            row['judge2_verdict'] = deepseek_data['verdict']
            row['judge2_score'] = deepseek_data['score']
            row['judge2_why'] = deepseek_data['why']
            updated_count += 1
    
    print(f"Updated {updated_count} rows with DeepSeek judge2 data")
    
    # Save updated per_item.csv
    fieldnames = [
        'run_id', 'config', 'topic', 'type', 'seed', 'source_text_sha', 'idx',
        'qid_gold', 'qid_pred', 'structural_valid',
        'question_sim', 'ans_sim', 'distractor_diversity', 'ans_score', 'item_score',
        'judge_verdict', 'judge_score', 'judge_why',
        'judge2_verdict', 'judge2_score', 'judge2_why'
    ]
    
    with per_item_csv.open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"Saved updated per_item.csv to {per_item_csv}")
    return updated_count


def main():
    import argparse
    ap = argparse.ArgumentParser(description='Restore DeepSeek judge2 data from report3.xlsx')
    ap.add_argument('--run-ids', nargs='+', default=['20251029_161124_gemini_deepseek', '20251029_161124_claude_deepseek'], 
                    help='Report IDs to restore DeepSeek data to')
    args = ap.parse_args()
    
    print("Loading DeepSeek judge2 data from report3.xlsx...")
    deepseek_map = load_deepseek_data_from_report3()
    
    for run_id in args.run_ids:
        print(f"\n{'=' * 70}")
        print(f"Restoring DeepSeek data to report: {run_id}")
        print(f"{'=' * 70}")
        try:
            updated = restore_deepseek_to_report(run_id, deepseek_map)
            print(f"Successfully restored {updated} DeepSeek judge2 entries")
        except Exception as e:
            print(f"Error: {e}")
    
    print(f"\n{'=' * 70}")
    print("Done! You may need to recompute summary.csv and other CSVs.")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    main()

