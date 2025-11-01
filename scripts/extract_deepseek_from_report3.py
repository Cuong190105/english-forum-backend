"""Extract DeepSeek judge2 data from report3.xlsx - check all sheets."""
from __future__ import annotations
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl


def find_deepseek_data_in_report3():
    """Find DeepSeek judge2 data in report3.xlsx by checking all sheets."""
    report3_path = Path("benchmark/reports/20251029_161124/report3.xlsx")
    
    if not report3_path.exists():
        raise FileNotFoundError(f"report3.xlsx not found: {report3_path}")
    
    wb = openpyxl.load_workbook(report3_path, data_only=True)
    
    print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    
    # Check each sheet for judge2 columns
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'=' * 70}")
        print(f"Checking sheet: {sheet_name}")
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        
        # Read first row for headers
        headers = []
        for col in range(1, min(ws.max_column + 1, 50)):
            cell_value = ws.cell(1, col).value
            if cell_value:
                headers.append((col, str(cell_value)))
        
        print(f"Headers (first 20): {[h[1] for h in headers[:20]]}")
        
        # Check if has judge2 columns
        has_judge2 = any('judge2' in str(h[1]).lower() for h in headers)
        has_per_item_cols = any(col in ['config', 'topic', 'type', 'seed', 'source_text_sha', 'idx'] 
                               for _, col in headers)
        
        if has_judge2:
            print(f"Found judge2 columns in sheet: {sheet_name}")
            judge2_cols = [h for h in headers if 'judge2' in str(h[1]).lower()]
            print(f"  Judge2 columns: {[h[1] for h in judge2_cols]}")
            
            # Check first few data rows
            if ws.max_row > 1:
                print(f"\nSample data from row 2:")
                row2_data = {}
                for col_idx, header in headers[:15]:
                    cell_value = ws.cell(2, col_idx).value
                    if cell_value:
                        row2_data[header] = str(cell_value)[:50]
                print(f"  {row2_data}")
        
        if has_per_item_cols and has_judge2:
            print(f"This sheet likely contains per_item data with judge2!")
            return sheet_name
    
    return None


if __name__ == '__main__':
    result = find_deepseek_data_in_report3()
    if result:
        print(f"\n{'=' * 70}")
        print(f"Found DeepSeek data in sheet: {result}")
    else:
        print(f"\n{'=' * 70}")
        print("Could not find DeepSeek judge2 data in report3.xlsx")
        print("You may need to check other files or restore from backup")

