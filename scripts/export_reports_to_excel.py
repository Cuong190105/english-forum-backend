from __future__ import annotations
import csv
import json
import hashlib
import re
import os
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
import warnings

# Suppress noisy NumPy runtime warnings early (before any imports that might pull NumPy)
os.environ.setdefault('PYTHONWARNINGS', 'ignore')
warnings.filterwarnings("ignore", category=RuntimeWarning)

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except Exception as e:
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl")


def safe_topic(topic: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", topic)


def load_per_item(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open('r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f, skipinitialspace=True)
        # Normalize header names by trimming whitespace
        if reader.fieldnames:
            reader.fieldnames = [(fn or '').strip() for fn in reader.fieldnames]
        for r in reader:
            if r is None:
                continue
            norm: Dict[str, Any] = {}
            for k, v in r.items():
                kk = (k or '').strip()
                if isinstance(v, str):
                    vv = v.strip()
                else:
                    vv = v
                norm[kk] = vv
            if not any(str(v).strip() for v in norm.values()):
                continue
            rows.append(norm)
    return rows


def load_summary(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open('r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f, skipinitialspace=True)
        if reader.fieldnames:
            reader.fieldnames = [(fn or '').strip() for fn in reader.fieldnames]
        for r in reader:
            if r is None:
                continue
            norm: Dict[str, Any] = {}
            for k, v in r.items():
                kk = (k or '').strip()
                if isinstance(v, str):
                    vv = v.strip()
                else:
                    vv = v
                norm[kk] = vv
            if not any(str(v).strip() for v in norm.values()):
                continue
            rows.append(norm)
    return rows


def load_input_map(jsonl_path: Optional[Path]) -> Dict[str, str]:
    """Map sha256(source_text) -> source_text for full text embedding in Excel.
    Accepts corrected_text, else source_text, else text.
    """
    if not jsonl_path or not jsonl_path.exists():
        return {}
    mp: Dict[str, str] = {}
    with jsonl_path.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            text = rec.get('corrected_text') or rec.get('source_text') or rec.get('text') or ''
            sha = hashlib.sha256(text.encode('utf-8')).hexdigest()
            mp[sha] = text
    return mp


def format_mcq_options(options: List[Dict[str, Any]]) -> str:
    # Produce a single-line compact options string
    parts = []
    for o in options:
        parts.append(f"{o.get('id')}) {o.get('label')}")
    return " | ".join(parts)


def build_excel(run_dir: Path, input_jsonl: Optional[Path] = None) -> Path:
    per_item_csv = run_dir / 'per_item.csv'
    summary_csv = run_dir / 'summary.csv'
    paired_overall_csv = run_dir / 'paired_overall.csv'
    winloss_csv = run_dir / 'winloss.csv'
    inter_judge_csv = run_dir / 'inter_judge.csv'
    inter_judge_by_topic_csv = run_dir / 'inter_judge_by_topic.csv'
    latency_csv = run_dir / 'latency.csv'
    if not per_item_csv.exists() or not summary_csv.exists():
        raise SystemExit(f"Missing per_item or summary CSV in {run_dir}")

    per_rows = load_per_item(per_item_csv)
    sum_rows = load_summary(summary_csv)
    text_map = load_input_map(input_jsonl)
    print(f"Loaded per_item rows: {len(per_rows)}; summary rows: {len(sum_rows)}; source_text mapped: {len(text_map)}")

    # Cache gold/pred items by per-source paths using source_text_sha
    gold_cache: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}
    pred_cache: Dict[Tuple[str, str, str, str, str], List[Dict[str, Any]]] = {}

    def get_gold(topic: str, qtype: str, sha: Optional[str]) -> List[Dict[str, Any]]:
        st = safe_topic(topic)
        key = (st, qtype, sha or '')
        if key in gold_cache:
            return gold_cache[key]
        if sha:
            p = Path(f"benchmark/gold/{st}/{qtype}/{sha}/seed0.json")
        else:
            # Fallback to old flat path (legacy runs)
            p = Path(f"benchmark/gold/{st}/{qtype}/seed0.json")
        items = json.loads(p.read_text(encoding='utf-8')) if p.exists() else []
        gold_cache[key] = items
        return items

    def get_pred(config: str, seed: str, topic: str, qtype: str, sha: Optional[str]) -> List[Dict[str, Any]]:
        st = safe_topic(topic)
        key = (config, seed, st, qtype, sha or '')
        if key in pred_cache:
            return pred_cache[key]
        if sha:
            p = Path(f"benchmark/pred/{config}/{st}/{qtype}/{sha}/seed{seed}.json")
        else:
            # Fallback to old flat path (legacy runs)
            p = Path(f"benchmark/pred/{config}/{st}/{qtype}/seed{seed}.json")
        items = json.loads(p.read_text(encoding='utf-8')) if p.exists() else []
        pred_cache[key] = items
        return items

    wb = Workbook()
    ws_sum: Worksheet = wb.active
    ws_sum.title = 'summary'
    ws_cmp: Worksheet = wb.create_sheet('compare')

    # Write summary sheet
    sum_cols = [
        'run_id','config','topic','type','seed','n_items','structural_pass_pct',
        'semantic_mean','semantic_std','semantic_se','semantic_ci95_low','semantic_ci95_high',
        'judge_mean','judge_std','judge_se','judge_ci95_low','judge_ci95_high',
        # Optional judge2 aggregate metrics if present in CSV
        'judge2_mean','judge2_std','judge2_se','judge2_ci95_low','judge2_ci95_high'
    ]
    ws_sum.append(sum_cols)
    for r in sum_rows:
        ws_sum.append([r.get(c, '') for c in sum_cols])

    # Write compare sheet
    cmp_cols = [
        'run_id','config','topic','type','seed','idx',
        'source_text',
        'gold_question','gold_answer_or_correct','gold_options',
        'pred_question','pred_answer_or_correct','pred_options',
        'structural_valid','question_sim','ans_sim','distractor_diversity','ans_score','item_score',
        'judge_verdict','judge_score','judge_why',
        # Optional judge2 per-item details
        'judge2_verdict','judge2_score','judge2_why'
    ]
    ws_cmp.append(cmp_cols)

    # Quick preflight for missing gold/pred files (per-source aware)
    missing_gold = set()
    missing_pred = set()
    def _get(r: Dict[str, Any], k: str) -> str:
        return str(r.get(k, '')).strip()
    for r in per_rows:
        topic = _get(r, 'topic')
        qtype = _get(r, 'type')
        config = _get(r, 'config')
        seed = _get(r, 'seed')
        st = safe_topic(topic)
        sha = _get(r, 'source_text_sha') or None
        if sha:
            gp = Path(f"benchmark/gold/{st}/{qtype}/{sha}/seed0.json")
            pp = Path(f"benchmark/pred/{config}/{st}/{qtype}/{sha}/seed{seed}.json")
        else:
            gp = Path(f"benchmark/gold/{st}/{qtype}/seed0.json")
            pp = Path(f"benchmark/pred/{config}/{st}/{qtype}/seed{seed}.json")
        if not gp.exists():
            missing_gold.add((topic, qtype, sha or ''))
        if not pp.exists():
            missing_pred.add((config, seed, topic, qtype, sha or ''))
    if missing_gold:
        print(f"Note: missing gold for {len(missing_gold)} topic/type combos — gold columns may be blank.")
    if missing_pred:
        print(f"Note: missing pred for {len(missing_pred)} config/seed/topic/type combos — pred columns may be blank.")

    for r in per_rows:
        topic = _get(r, 'topic')
        qtype = _get(r, 'type')
        config = _get(r, 'config')
        seed = _get(r, 'seed')
        idx_s = _get(r, 'idx')
        try:
            idx = int(idx_s)
        except Exception:
            # Skip rows without a valid idx
            continue
        source_text = ''
        sha = _get(r, 'source_text_sha') or None
        if sha and sha in text_map:
            source_text = text_map[sha]

        gold_items = get_gold(topic, qtype, sha)
        pred_items = get_pred(config, seed, topic, qtype, sha)
        gi = gold_items[idx-1] if idx-1 < len(gold_items) else {}
        pi = pred_items[idx-1] if idx-1 < len(pred_items) else {}

        if qtype == 'mcq':
            g_prompt = gi.get('question',{}).get('prompt','')
            g_cid = gi.get('correctOptionId','')
            g_opts = gi.get('question',{}).get('options',[])
            g_corr = next((o.get('label') for o in g_opts if o.get('id')==g_cid), '')
            g_opts_s = format_mcq_options(g_opts)

            p_prompt = pi.get('question',{}).get('prompt','')
            p_cid = pi.get('correctOptionId','')
            p_opts = pi.get('question',{}).get('options',[])
            p_corr = next((o.get('label') for o in p_opts if o.get('id')==p_cid), '')
            p_opts_s = format_mcq_options(p_opts)

            row = [
                _get(r,'run_id'), config, topic, qtype, _get(r,'seed'), idx,
                source_text,
                g_prompt, g_corr, g_opts_s,
                p_prompt, p_corr, p_opts_s,
                _get(r,'structural_valid'), (_get(r,'question_sim') or _get(r,'prompt_sim')), _get(r,'ans_sim'), _get(r,'distractor_diversity'), _get(r,'ans_score'), _get(r,'item_score'), _get(r,'judge_verdict'), _get(r,'judge_score'), _get(r,'judge_why'), _get(r,'judge2_verdict'), _get(r,'judge2_score'), _get(r,'judge2_why')
            ]
        else:
            g_prompt = gi.get('question',{}).get('prompt','')
            g_ans = gi.get('answer','')

            p_prompt = pi.get('question',{}).get('prompt','')
            p_ans = pi.get('answer','')

            row = [
                _get(r,'run_id'), config, topic, qtype, _get(r,'seed'), idx,
                source_text,
                g_prompt, g_ans, '',
                p_prompt, p_ans, '',
                _get(r,'structural_valid'), (_get(r,'question_sim') or _get(r,'prompt_sim')), _get(r,'ans_sim'), '', _get(r,'ans_score'), _get(r,'item_score'), _get(r,'judge_verdict'), _get(r,'judge_score'), _get(r,'judge_why'), _get(r,'judge2_verdict'), _get(r,'judge2_score'), _get(r,'judge2_why')
            ]

        ws_cmp.append(row)

    # Optional sheets: paired_overall and winloss
    def add_sheet_from_csv(csv_path: Path, title: str):
        if not csv_path.exists():
            print(f"Note: {csv_path.name} not found; skipping '{title}' sheet.")
            return None
        with csv_path.open('r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            print(f"Note: {csv_path.name} is empty; skipping '{title}' sheet.")
            return None
        ws: Worksheet = wb.create_sheet(title)
        for row in rows:
            ws.append(row)
        return ws

    ws_paired = add_sheet_from_csv(paired_overall_csv, 'paired_overall')
    ws_winloss = add_sheet_from_csv(winloss_csv, 'winloss')
    ws_inter = add_sheet_from_csv(inter_judge_csv, 'inter_judge')
    ws_inter_topic = add_sheet_from_csv(inter_judge_by_topic_csv, 'inter_judge_by_topic')
    
    # Add latency comparison sheet (minimal vs CoT)
    ws_latency = add_sheet_from_csv(latency_csv, 'latency')

    # Autosize columns for all sheets we created
    for ws in wb.worksheets:
        # Skip autosizing for extremely wide data to keep performance reasonable
        try:
            for col_idx, _ in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                col_letter = get_column_letter(col_idx)
                width = 0
                for cell in ws[col_letter]:
                    v = cell.value
                    if v is None:
                        continue
                    l = len(str(v))
                    if l > width:
                        width = l
                ws.column_dimensions[col_letter].width = min(max(width + 2, 10), 100)
        except Exception:
            # Best-effort sizing; ignore if sheet is unusual
            pass

    out_path = run_dir / 'report.xlsx'
    # If file exists and might be locked, try to save with a timestamped name
    if out_path.exists():
        try:
            # Try to save to the same file first
            wb.save(out_path)
        except PermissionError:
            # If locked, save with timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = run_dir / f'report_{timestamp}.xlsx'
            print(f"Warning: {out_path} is locked. Saving to {backup_path} instead.")
            wb.save(backup_path)
            out_path = backup_path
    else:
        wb.save(out_path)
    return out_path


def main():
    import argparse
    ap = argparse.ArgumentParser(description='Export benchmark CSVs to a human-friendly Excel with source text and gold vs pred side-by-side.')
    ap.add_argument('--run-id', help='Run ID (folder under benchmark/reports)')
    ap.add_argument('--run-dir', help='Path to a reports directory (if not using --run-id)')
    ap.add_argument('--input-jsonl', help='Optional labeled input JSONL to recover full source_text for display')
    ap.add_argument('--open-file', action='store_true', help='Open the generated Excel file after export')
    args = ap.parse_args()

    if not args.run_id and not args.run_dir:
        raise SystemExit('Please provide --run-id or --run-dir')

    run_dir = Path(args.run_dir) if args.run_dir else Path(f"benchmark/reports/{args.run_id}")
    input_jsonl = Path(args.input_jsonl) if args.input_jsonl else None
    out = build_excel(run_dir, input_jsonl)
    print(f"Excel written: {out}")
    if args.open_file:
        try:
            os.startfile(out)  # Windows-only convenience
        except Exception:
            pass


if __name__ == '__main__':
    main()
