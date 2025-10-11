"""Format the pretty Excel file for annotation: wrap text, freeze header, bold header, set column widths, autofilter.

Usage:
    python -m scripts.format_pretty_excel --input data/jfleg_eval_prefill_100_pretty.xlsx --output data/jfleg_eval_prefill_100_pretty_formatted.xlsx

Requires: openpyxl
"""
from __future__ import annotations
import argparse
import os
import sys
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


def auto_width(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ''
                l = len(v)
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        adjusted_width = min(max_width, max(10, int(max_len * 0.9)))
        ws.column_dimensions[col_letter].width = adjusted_width


def main(input_path: str, output_path: str):
    wb = load_workbook(input_path)
    ws = wb.active

    # Freeze header
    ws.freeze_panes = 'A2'

    # Bold header and wrap text
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Wrap text for columns likely to contain JSON
    headers = [c.value for c in ws[1]]
    wrap_keys = {
        'source_text',
        'cot_output', 'minimal_output',
        'cot_output_flash', 'minimal_output_flash',
        'cot_output_flash_lite', 'minimal_output_flash_lite',
    }
    for idx, h in enumerate(headers, start=1):
        key = (h or '').strip().lower()
        if key in wrap_keys or key.endswith('_output'):
            for rcell in ws[get_column_letter(idx)]:
                rcell.alignment = Alignment(wrap_text=True, vertical='top')

    # Set a reasonable default row height for better legibility on wrapped text
    ws.row_dimensions[1].height = 24
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 90

    # Autofilter on header
    ws.auto_filter.ref = ws.dimensions

    # Set reasonable column widths
    auto_width(ws)

    wb.save(output_path)
    print(f"Wrote formatted Excel to {output_path}")


if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--input', default='data/jfleg_eval_prefill_100_pretty.xlsx')
    p.add_argument('--output', default='data/jfleg_eval_prefill_100_pretty_formatted.xlsx')
    args = p.parse_args()
    main(args.input, args.output)
